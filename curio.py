from openpyxl import load_workbook

class Cabinet:

    def __init__(self, directory):
        self.directory = directory
        self._parse_spreadsheet()

    def _parse_spreadsheet(self):
        workbook = load_workbook(filename="test-data/conrad.xlsx")
        worksheet = workbook[workbook.sheetnames[0]]
        self.items = []
        for row in worksheet.rows:
            self.items.append(row[1])

