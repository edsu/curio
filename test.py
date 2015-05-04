#!/usr/bin/env python

from openpyxl import load_workbook

workbook = load_workbook(filename="test-data/conrad.xlsx")
worksheet = workbook[workbook.sheetnames[0]]

for row in worksheet.rows:
    print(row[1].value)
